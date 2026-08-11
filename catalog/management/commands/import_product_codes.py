"""Проставить код товара (sku) и штрихкод из файла ценников .xls/.xlsx."""

from __future__ import annotations

import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from catalog.models import Product


def _cell_str(value) -> str:
    if value is None:
        return ''
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).rstrip('0').rstrip('.')
    return str(value).strip()


def _norm_name(value: str) -> str:
    s = (value or '').strip().lower().replace('ё', 'е')
    s = s.replace('(', ' ').replace(')', ' ')
    s = s.replace('л.', 'л').replace('тр.', 'тр')
    s = re.sub(r'[\s\u00a0]+', ' ', s)
    s = re.sub(r'[.,]', ',', s)
    return s.strip()


def _norm_compact(value: str) -> str:
    return re.sub(r'[^a-z0-9а-я]+', '', _norm_name(value))


def _read_rows(path: Path):
    suffix = path.suffix.lower()
    rows = []
    if suffix == '.xls':
        import xlrd

        book = xlrd.open_workbook(str(path))
        sheet = book.sheet_by_index(0)
        headers = [_cell_str(sheet.cell_value(0, c)).lower() for c in range(sheet.ncols)]
        for r in range(1, sheet.nrows):
            rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
        return headers, rows

    if suffix in {'.xlsx', '.xlsm'}:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), read_only=True, data_only=True)
        sheet = wb.active
        data = list(sheet.iter_rows(values_only=True))
        if not data:
            return [], []
        headers = [_cell_str(v).lower() for v in data[0]]
        rows = [list(row) for row in data[1:]]
        return headers, rows

    raise CommandError('Поддерживаются только .xls / .xlsx')


def _col(headers, *aliases):
    for alias in aliases:
        alias = alias.lower()
        for idx, header in enumerate(headers):
            if alias in header:
                return idx
    return None


class Command(BaseCommand):
    help = 'Импорт кодов товара (sku) и штрихкодов из файла ценников'

    def add_arguments(self, parser):
        parser.add_argument('path', type=str, help='Путь к .xls/.xlsx')
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Только показать результат, без записи в БД',
        )

    @transaction.atomic
    def handle(self, *args, **options):
        path = Path(options['path']).expanduser().resolve()
        if not path.is_file():
            raise CommandError(f'Файл не найден: {path}')

        headers, rows = _read_rows(path)
        name_i = _col(headers, 'название товара', 'название', 'наименование')
        sku_i = _col(headers, 'код товара', 'артикул', 'код')
        barcode_i = _col(headers, 'штрихкод', 'штрих-код', 'barcode', 'ean')
        if name_i is None or sku_i is None:
            raise CommandError(f'Не найдены колонки названия/кода. Заголовки: {headers}')

        by_name = {}
        by_compact = {}
        by_barcode = {}
        for p in Product.objects.all():
            key = _norm_name(p.name)
            by_name.setdefault(key, []).append(p)
            by_compact.setdefault(_norm_compact(p.name), []).append(p)
            if p.barcode:
                by_barcode[_cell_str(p.barcode)] = p

        updated = 0
        skipped_no_code = 0
        unmatched = []

        for row in rows:
            if name_i >= len(row):
                continue
            name = _cell_str(row[name_i])
            sku = _cell_str(row[sku_i]) if sku_i < len(row) else ''
            barcode = ''
            if barcode_i is not None and barcode_i < len(row):
                barcode = _cell_str(row[barcode_i])

            if not name:
                continue
            if not sku:
                skipped_no_code += 1
                continue
            # В файле иногда в колонку кода попадает цена вида 150,00
            if re.fullmatch(r'\d+[,.]\d{2}', sku):
                skipped_no_code += 1
                continue

            product = None
            if barcode and barcode in by_barcode:
                product = by_barcode[barcode]
            if not product:
                key = _norm_name(name)
                candidates = by_name.get(key) or by_compact.get(_norm_compact(name)) or []
                if len(candidates) == 1:
                    product = candidates[0]
                elif len(candidates) > 1 and barcode:
                    for cand in candidates:
                        if _cell_str(cand.barcode) == barcode:
                            product = cand
                            break
                    # Одинаковые названия без штрихкода в БД — раздаём по порядку
                    if not product:
                        free = [c for c in candidates if not (c.sku or '').strip()]
                        if free:
                            product = free[0]

            if not product:
                unmatched.append((name, sku))
                continue

            fields = []
            if product.sku != sku:
                product.sku = sku[:100]
                fields.append('sku')
            if barcode and product.barcode != barcode:
                product.barcode = barcode[:64]
                fields.append('barcode')
            if fields and not options['dry_run']:
                product.save(update_fields=fields)
            if fields:
                updated += 1

        if options['dry_run']:
            transaction.set_rollback(True)

        self.stdout.write(
            self.style.SUCCESS(
                f'Обновлено: {updated}; без кода в файле: {skipped_no_code}; '
                f'не сопоставлено: {len(unmatched)}'
            )
        )
        for name, sku in unmatched[:40]:
            self.stdout.write(f'  ? {sku} — {name}')
        if len(unmatched) > 40:
            self.stdout.write(f'  … ещё {len(unmatched) - 40}')
