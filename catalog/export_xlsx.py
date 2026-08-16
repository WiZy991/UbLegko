"""Выгрузка публичного каталога в Excel с фотографиями."""

from __future__ import annotations

from io import BytesIO
from itertools import groupby
from pathlib import Path

from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PILImage, ImageOps

from .models import Product

COL_WIDTHS = {
    'A': 16,
    'B': 14,
    'C': 18,   # Название
    'D': 105,  # Описание
    'E': 10,
    'F': 12,
    'G': 10,
    'H': 14,
}
NUM_COLS = 8
ROW_HEIGHT_MIN = 22
ROW_HEIGHT_MIN_WITH_PHOTO = 52
ROW_HEIGHT_MAX = 280
LINE_HEIGHT_PT = 13.5
ROW_PADDING_PT = 10
CATEGORY_ROW_HEIGHT = 26
PHOTO_THUMB = (78, 78)
BADGE_THUMB = (40, 40)
# Картинка занимает не больше этой доли ячейки — иначе Excel рисует поверх границ
MAX_CELL_FILL = 0.58


def catalog_xlsx_filename() -> str:
    """Имя файла для скачивания (без двоеточия — Windows его запрещает)."""
    today = timezone.localdate().strftime('%d.%m.%y')
    return f'Прайс магазина Убираемся легко по состоянию на {today}.xlsx'


def _col_width_px(excel_width: float) -> int:
    return max(1, int(excel_width * 7 + 5))


def _row_height_px(points: float) -> int:
    return max(1, int(points * 96 / 72))


def _wrapped_line_count(text: str, excel_width: float) -> int:
    """Грубая оценка числа строк при wrap_text в Excel."""
    text = (text or '').replace('\r\n', '\n').replace('\r', '\n')
    if not text.strip():
        return 1
    # Ширина колонки ≈ число символов «0»; кириллица чуть шире
    chars_per_line = max(4.0, excel_width * 0.92)
    total = 0
    for paragraph in text.split('\n'):
        if paragraph == '':
            total += 1
            continue
        total += max(1, int((len(paragraph) + chars_per_line - 1) // chars_per_line))
    return max(1, total)


def _row_height_for_content(
    name: str,
    description: str,
    *,
    has_photo: bool,
) -> float:
    """Высота строки по названию/описанию: мало текста — низкая, много — высокая."""
    lines = max(
        _wrapped_line_count(name, COL_WIDTHS['C']),
        _wrapped_line_count(description, COL_WIDTHS['D']),
        1,
    )
    height = lines * LINE_HEIGHT_PT + ROW_PADDING_PT
    min_h = ROW_HEIGHT_MIN_WITH_PHOTO if has_photo else ROW_HEIGHT_MIN
    return max(min_h, min(ROW_HEIGHT_MAX, height))


def _badge_akciya_path() -> Path | None:
    candidates = [
        Path(settings.BASE_DIR) / 'static' / 'img' / 'badge-akciya.png',
        Path(settings.STATIC_ROOT) / 'img' / 'badge-akciya.png' if settings.STATIC_ROOT else None,
    ]
    for path in candidates:
        if path and path.is_file():
            return path
    return None


def _product_image_path(product: Product) -> Path | None:
    """Для Excel берём лёгкое превью каталога, иначе оригинал."""
    for field_name in ('image_card', 'image'):
        field = getattr(product, field_name, None)
        if not field:
            continue
        try:
            path = Path(field.path)
        except Exception:
            continue
        if path.is_file():
            return path
    for item in product.images.all():
        if not item.image:
            continue
        try:
            path = Path(item.image.path)
        except Exception:
            continue
        if path.is_file():
            return path
    return None


def _thumb_bytes(path: Path, thumb_size: tuple[int, int]) -> tuple[BytesIO, int, int] | None:
    """Уменьшенная картинка для ячейки Excel (JPEG — быстрее PNG)."""
    try:
        with PILImage.open(path) as img:
            img = ImageOps.exif_transpose(img)
            if img.mode in ('RGBA', 'LA') or (img.mode == 'P' and 'transparency' in img.info):
                rgba = img.convert('RGBA')
                background = PILImage.new('RGB', rgba.size, (255, 255, 255))
                background.paste(rgba, mask=rgba.split()[-1])
                img = background
            else:
                img = img.convert('RGB')
            img.thumbnail(thumb_size, PILImage.Resampling.BILINEAR)
            buf = BytesIO()
            img.save(buf, format='JPEG', quality=72, optimize=True)
            buf.seek(0)
            return buf, img.width, img.height
    except Exception:
        return None


def _add_image_centered_in_cell(
    ws,
    buf: BytesIO,
    *,
    col_idx: int,
    row_idx: int,
    col_width_excel: float,
    row_height_pt: float,
    img_w: int,
    img_h: int,
) -> None:
    """Ставит картинку по центру ячейки с запасом от границ (не перекрывает линии)."""
    cell_w = _col_width_px(col_width_excel)
    cell_h = _row_height_px(row_height_pt)

    max_w = max(12, int(cell_w * MAX_CELL_FILL))
    max_h = max(12, int(cell_h * MAX_CELL_FILL))
    scale = min(max_w / img_w, max_h / img_h, 1.0)
    disp_w = max(1, int(img_w * scale))
    disp_h = max(1, int(img_h * scale))

    offset_x = max(0, (cell_w - disp_w) // 2)
    offset_y = max(0, (cell_h - disp_h) // 2)

    img = XLImage(buf)
    img.width = disp_w
    img.height = disp_h
    marker = AnchorMarker(
        col=col_idx - 1,
        colOff=pixels_to_EMU(offset_x),
        row=row_idx - 1,
        rowOff=pixels_to_EMU(offset_y),
    )
    size = XDRPositiveSize2D(pixels_to_EMU(disp_w), pixels_to_EMU(disp_h))
    img.anchor = OneCellAnchor(_from=marker, ext=size)
    ws.add_image(img)


def _write_category_banner(ws, row_idx: int, title: str, *, thin: Border) -> None:
    """Горизонтальная полоса на всю ширину таблицы с названием категории."""
    category_font = Font(bold=True, color='FFFFFF', size=13)
    category_fill = PatternFill('solid', fgColor='0A4F61')
    category_align = Alignment(horizontal='left', vertical='center', indent=1)

    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=NUM_COLS)
    cell = ws.cell(row_idx, 1, title)
    cell.font = category_font
    cell.fill = category_fill
    cell.alignment = category_align
    cell.border = thin

    for col in range(2, NUM_COLS + 1):
        side = ws.cell(row_idx, col)
        side.fill = category_fill
        side.border = thin

    ws.row_dimensions[row_idx].height = CATEGORY_ROW_HEIGHT


def _write_product_row(
    ws,
    row_idx: int,
    product: Product,
    *,
    thin: Border,
    wrap: Alignment,
    center: Alignment,
    image_buffers: list[BytesIO],
    badge_cache: tuple[bytes, int, int] | None,
) -> None:
    is_promo = bool(product.is_promo or (product.old_price and product.old_price > 0))
    path = _product_image_path(product)
    name = product.name or ''
    description = product.description or ''
    row_height = _row_height_for_content(name, description, has_photo=bool(path))

    if path:
        result = _thumb_bytes(path, PHOTO_THUMB)
        if result:
            buf, iw, ih = result
            image_buffers.append(buf)
            _add_image_centered_in_cell(
                ws,
                buf,
                col_idx=1,
                row_idx=row_idx,
                col_width_excel=COL_WIDTHS['A'],
                row_height_pt=row_height,
                img_w=iw,
                img_h=ih,
            )

    if is_promo and badge_cache:
        raw, iw, ih = badge_cache
        buf = BytesIO(raw)
        image_buffers.append(buf)
        _add_image_centered_in_cell(
            ws,
            buf,
            col_idx=6,
            row_idx=row_idx,
            col_width_excel=COL_WIDTHS['F'],
            row_height_pt=row_height,
            img_w=iw,
            img_h=ih,
        )

    values = [
        '',
        product.sku or '',
        name,
        description,
        float(product.price) if product.price is not None else '',
        '',
        float(product.rating) if product.rating is not None else '',
        product.country or '',
    ]
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row_idx, col, value)
        cell.border = thin
        cell.alignment = center if col in (1, 2, 5, 6, 7, 8) else wrap

    ws.row_dimensions[row_idx].height = row_height


def build_catalog_xlsx(*, site_origin: str = '') -> bytes:
    products = list(
        Product.objects.filter(is_visible=True)
        .select_related('category')
        .prefetch_related('images')
        .order_by('category__sort_order', 'category__name', 'name')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Прайс'

    headers = [
        'Фото',
        'Артикул',
        'Название',
        'Описание',
        'Цена',
        'Акция',
        'Рейтинг',
        'Страна',
    ]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='0F6F86')
    thin = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    wrap = Alignment(wrap_text=True, vertical='center')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col, title in enumerate(headers, start=1):
        cell = ws.cell(1, col, title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin

    for letter, width in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    image_buffers: list[BytesIO] = []
    badge_path = _badge_akciya_path()
    badge_cache: tuple[bytes, int, int] | None = None
    if badge_path:
        badge_result = _thumb_bytes(badge_path, BADGE_THUMB)
        if badge_result:
            buf, iw, ih = badge_result
            badge_cache = (buf.getvalue(), iw, ih)
    row_idx = 2

    for category_key, group in groupby(
        products,
        key=lambda p: (
            p.category_id,
            p.category.name if p.category_id else 'Без категории',
        ),
    ):
        _category_id, category_name = category_key
        category_products = list(group)
        if not category_products:
            continue

        _write_category_banner(ws, row_idx, category_name, thin=thin)
        row_idx += 1

        for product in category_products:
            _write_product_row(
                ws,
                row_idx,
                product,
                thin=thin,
                wrap=wrap,
                center=center,
                image_buffers=image_buffers,
                badge_cache=badge_cache,
            )
            row_idx += 1

    ws.oddHeader.center.text = (
        f'Прайс магазина Убираемся легко по состоянию на '
        f'{timezone.localdate().strftime("%d.%m.%y")} · товаров: {len(products)}'
    )

    out = BytesIO()
    wb.save(out)
    _ = image_buffers
    return out.getvalue()
