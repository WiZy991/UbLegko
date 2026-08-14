import csv
import io
import json
import logging
import re
from decimal import Decimal, InvalidOperation

from django.contrib import admin, messages
from django.contrib.admin.templatetags.admin_list import (
    ResultList,
    items_for_result,
    result_headers,
    result_hidden_fields,
)
from django.core.files.base import ContentFile
from django.db.models import Count
from django.http import JsonResponse
from django.shortcuts import redirect, render
from django.urls import path, reverse
from django.utils.html import escape, format_html, format_html_join, mark_safe
from openpyxl import load_workbook
from pathlib import Path

from .categorize import resolve_category
from .models import Category, Product, ProductImage, ProductRecommendation, ProductReview
from .search_utils import filter_products_by_query, rank_prefix_first

logger = logging.getLogger(__name__)

HEADER_ALIASES = {
    'name': {
        'name',
        'название',
        'наименование',
        'наименование товара',
        'товар',
    },
    'description': {
        'description',
        'описание',
        'описание товара',
    },
    # Только для импорта старых прайсов — в модель не пишется
    'short_description': {
        'short_description',
        'краткое описание',
    },
    'price': {
        'price',
        'цена',
        'цена руб',
        'цена, руб',
    },
    'old_price': {
        'old_price',
        'старая цена',
        'цена старая',
    },
    'country': {
        'country',
        'страна',
        'страна производитель',
        'страна-производитель',
        'страна-произво-дитель',
        'производитель',
    },
    'sku': {
        'sku',
        'код',
        'код товара',
        'артикул',
    },
    'barcode': {
        'barcode',
        'штрихкод',
        'штрих-код',
        'ean',
    },
    'category': {
        'category',
        'категория',
        'раздел',
    },
    'status': {'status', 'статус'},
    'is_promo': {'is_promo', 'акция', 'promo'},
}


def normalize_header(value):
    text = str(value or '').strip().lower().replace('\n', ' ')
    text = re.sub(r'\s+', ' ', text)
    return text


def map_headers(raw_headers):
    mapping = {}
    for idx, header in enumerate(raw_headers):
        normalized = normalize_header(header)
        for field, aliases in HEADER_ALIASES.items():
            if normalized in aliases and field not in mapping:
                mapping[field] = idx
                break
    return mapping


def parse_price(value):
    if value is None or value == '':
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace(' ', '').replace('руб', '').replace('₽', '')
    text = text.replace(',', '.')
    text = re.sub(r'[^0-9.\-]', '', text)
    if not text:
        return None
    return Decimal(text)


class ProductImageInline(admin.TabularInline):
    model = ProductImage
    extra = 1
    fields = ('image', 'sort_order', 'preview')
    readonly_fields = ('preview',)
    ordering = ('sort_order', 'id')
    verbose_name = 'Фото'
    verbose_name_plural = 'Дополнительные фотографии (можно добавить сколько угодно)'

    def preview(self, obj):
        if obj.pk and obj.image:
            return format_html(
                '<img src="{}" style="height:64px;width:64px;object-fit:cover;border-radius:6px;" />',
                obj.image.url,
            )
        return '—'

    preview.short_description = 'Превью'


class ProductRecommendationInline(admin.TabularInline):
    model = ProductRecommendation
    fk_name = 'product'
    extra = 1
    autocomplete_fields = ['recommended_product']
    verbose_name = 'Рекомендуем к этому товару'
    verbose_name_plural = 'Рекомендуем к этому товару'


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ('drag_handle', 'name', 'is_visible', 'products_count')
    list_display_links = ('name',)
    list_editable = ('is_visible',)
    list_filter = ('is_visible',)
    search_fields = ('name',)
    ordering = ('sort_order', 'name')
    prepopulated_fields = {'slug': ('name',)}
    change_list_template = 'admin/catalog/category/change_list.html'

    class Media:
        css = {'all': ('admin/css/category_sortable.css',)}

    @admin.display(description='')
    def drag_handle(self, obj):
        return mark_safe(
            '<span class="category-drag-handle" title="Перетащите для изменения порядка" aria-hidden="true">⠿</span>'
        )

    @admin.display(description='Товаров', ordering='products_count')
    def products_count(self, obj):
        return getattr(obj, 'products_count', obj.products.count())

    def get_queryset(self, request):
        return super().get_queryset(request).annotate(products_count=Count('products'))

    def get_search_results(self, request, queryset, search_term):
        """
        Подстрочный поиск без учёта регистра (и для кириллицы).
        «руч» → «Для ручной мойки посуды».
        SQLite LIKE/icontains регистронезависим только для латиницы.
        """
        term = (search_term or '').strip()
        if not term:
            return queryset, False
        needle = term.replace('ё', 'е').replace('Ё', 'е').casefold()
        matching_ids = [
            pk
            for pk, name in queryset.values_list('pk', 'name')
            if needle in (name or '').replace('ё', 'е').replace('Ё', 'е').casefold()
        ]
        return queryset.filter(pk__in=matching_ids).order_by('sort_order', 'name'), False

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                'reorder/',
                self.admin_site.admin_view(self.reorder_view),
                name='catalog_category_reorder',
            ),
        ]
        return custom + urls

    def reorder_view(self, request):
        if request.method != 'POST':
            return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)
        if not self.has_change_permission(request):
            return JsonResponse({'ok': False, 'error': 'Forbidden'}, status=403)
        try:
            payload = json.loads(request.body.decode('utf-8') or '{}')
            order = payload.get('order') or []
        except (TypeError, ValueError, json.JSONDecodeError):
            return JsonResponse({'ok': False, 'error': 'Bad JSON'}, status=400)

        ids = []
        for raw in order:
            try:
                ids.append(int(raw))
            except (TypeError, ValueError):
                continue

        existing = set(
            Category.objects.filter(pk__in=ids).values_list('pk', flat=True)
        )
        updates = []
        for index, pk in enumerate(ids):
            if pk not in existing:
                continue
            updates.append(Category(pk=pk, sort_order=index))
        if updates:
            Category.objects.bulk_update(updates, ['sort_order'])
        return JsonResponse({'ok': True, 'count': len(updates)})


PRODUCT_ADMIN_SORT = {
    'category': ('category__sort_order', 'category__name', 'name'),
    'alpha': ('name',),
    'price_asc': ('price', 'name'),
    'price_desc': ('-price', 'name'),
    'promo': ('category__sort_order', 'category__name', 'name'),
}


class ProductSortFilter(admin.SimpleListFilter):
    title = 'Сортировка'
    parameter_name = 'sort'

    def lookups(self, request, model_admin):
        return (
            ('category', 'По категориям'),
            ('alpha', 'По алфавиту'),
            ('price_asc', 'Сначала дешевые'),
            ('price_desc', 'Сначала дорогие'),
            ('promo', 'Акции'),
        )

    def queryset(self, request, queryset):
        if self.value() == 'promo':
            return queryset.filter(is_promo=True)
        return queryset


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = (
        'expand_toggle',
        'name',
        'category',
        'price',
        'old_price',
        'status',
        'recommendation_codes',
        'row_save',
        'is_visible',
        'has_main_photo',
    )
    # name в list_editable — нельзя держать в list_display_links
    list_display_links = None
    list_filter = (ProductSortFilter, 'category', 'status', 'is_promo', 'is_visible', 'is_featured', 'country')
    list_editable = (
        'name',
        'category',
        'price',
        'old_price',
        'status',
        'recommendation_codes',
        'is_visible',
    )
    list_select_related = ('category',)
    list_per_page = 50
    search_fields = ('name', 'sku', 'barcode', 'description', 'country', 'recommendation_codes')
    prepopulated_fields = {'slug': ('name',)}
    autocomplete_fields = ['category']
    inlines = [ProductImageInline]
    change_list_template = 'admin/catalog/product/change_list.html'
    readonly_fields = ('rating', 'reviews_count')
    fieldsets = (
        (None, {
            'fields': (
                'name', 'slug', 'category', 'sku', 'barcode',
                'description',
                'unit', 'country', 'image',
            ),
        }),
        ('Цены и статус', {
            'fields': (
                'price',
                'old_price',
                'status',
                'recommendation_codes',
                'rating',
                'reviews_count',
                'is_promo',
                'is_featured',
                'is_visible',
            ),
            'description': (
                'Если указана «Старая цена», товар автоматически становится акционным '
                'и на сайте показывается бэйдж «Акция». '
                'Если «Старую цену» очистить — товар снимается с акции и бэйдж пропадает. '
                'В «Цена» — текущая цена, в «Старая цена» — прежняя (зачёркнутая). '
                '«Рекомендация» — номера групп через запятую (1 или 1,3,5): товары с общим '
                'номером показываются друг другу в блоке рекомендаций.'
            ),
        }),
    )

    def get_queryset(self, request):
        # Галерею не prefetch'им на весь список — фото подгружаются при раскрытии строки
        return super().get_queryset(request).select_related('category')

    def get_ordering(self, request):
        sort = request.GET.get('sort')
        if sort in PRODUCT_ADMIN_SORT:
            return PRODUCT_ADMIN_SORT[sort]
        return ('name',)

    def get_search_results(self, request, queryset, search_term):
        """Тот же поиск, что на сайте: название, описание, артикул, раскладка, стемминг."""
        search_term = (search_term or '').strip()
        if not search_term:
            return queryset, False
        return filter_products_by_query(queryset, search_term, prefix_only=False), False

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['product_search_suggest_url'] = reverse(
            'admin:catalog_product_search_suggest'
        )
        response = super().changelist_view(request, extra_context=extra_context)
        if not hasattr(response, 'context_data'):
            return response

        cl = response.context_data.get('cl')
        if not cl:
            return response

        rows = []
        if cl.formset:
            for res, form in zip(cl.result_list, cl.formset.forms):
                rows.append({
                    'result': ResultList(form, items_for_result(cl, res, form)),
                    'obj': res,
                    'details_url': reverse(
                        'admin:catalog_product_quick_details', args=[res.pk]
                    ),
                })
        else:
            for res in cl.result_list:
                rows.append({
                    'result': ResultList(None, items_for_result(cl, res, None)),
                    'obj': res,
                    'details_url': reverse(
                        'admin:catalog_product_quick_details', args=[res.pk]
                    ),
                })

        headers = list(result_headers(cl))
        num_sorted_fields = sum(
            1 for header in headers if header.get('sortable') and header.get('sorted')
        )
        response.context_data.update({
            'expandable_rows': rows,
            'result_headers': headers,
            'num_sorted_fields': num_sorted_fields,
            'result_hidden_fields': list(result_hidden_fields(cl)),
        })
        return response

    @admin.display(description='')
    def expand_toggle(self, obj):
        return format_html(
            '<button type="button" class="product-row-expand" data-product-id="{}" '
            'aria-expanded="false" aria-label="Подробнее о товаре" title="Подробнее">'
            '<i class="fas fa-chevron-down" aria-hidden="true"></i>'
            '</button>',
            obj.pk,
        )

    @admin.display(description='Сохранить')
    def row_save(self, obj):
        url = reverse('admin:catalog_product_quick_update', args=[obj.pk])
        return format_html(
            '<button type="button" class="product-row-save" data-product-save="{0}" '
            'data-quick-url="{1}" disabled title="Сохранить изменения товара">'
            'Сохранить'
            '</button>',
            obj.pk,
            url,
        )

    def photos_html(self, obj):
        photos_url = reverse('admin:catalog_product_quick_photos', args=[obj.pk])
        cards = []

        if obj.image:
            cards.append(
                format_html(
                    '<div class="product-row-detail__photo" data-photo-kind="main">'
                    '<img src="{}" class="product-row-detail__thumb" alt="">'
                    '<span class="product-row-detail__photo-badge">Главное</span>'
                    '<div class="product-row-detail__photo-actions">'
                    '<label class="button product-row-detail__file-btn">'
                    'Заменить'
                    '<input type="file" accept="image/*" hidden data-photo-upload="main">'
                    '</label>'
                    '<button type="button" class="button" data-photo-delete="main">Удалить</button>'
                    '</div>'
                    '</div>',
                    obj.image.url,
                )
            )
        else:
            cards.append(
                mark_safe(
                    '<div class="product-row-detail__photo product-row-detail__photo--empty" data-photo-kind="main">'
                    '<div class="product-row-detail__photo-empty">Нет главного фото</div>'
                    '<div class="product-row-detail__photo-actions">'
                    '<label class="button product-row-detail__file-btn">'
                    'Загрузить главное'
                    '<input type="file" accept="image/*" hidden data-photo-upload="main">'
                    '</label>'
                    '</div>'
                    '</div>'
                )
            )

        for image in obj.images.all():
            if not image.image:
                continue
            cards.append(
                format_html(
                    '<div class="product-row-detail__photo" data-photo-kind="gallery" data-photo-id="{}">'
                    '<img src="{}" class="product-row-detail__thumb" alt="">'
                    '<span class="product-row-detail__photo-badge">Галерея</span>'
                    '<div class="product-row-detail__photo-actions">'
                    '<button type="button" class="button" data-photo-set-main="{}">Сделать главным</button>'
                    '<button type="button" class="button" data-photo-delete="gallery" data-photo-id="{}">Удалить</button>'
                    '</div>'
                    '</div>',
                    image.pk,
                    image.image.url,
                    image.pk,
                    image.pk,
                )
            )

        cards.append(
            mark_safe(
                '<label class="product-row-detail__photo product-row-detail__photo--add">'
                '<span class="product-row-detail__photo-add-text">+ Добавить фото</span>'
                '<span class="product-row-detail__photo-add-hint">можно несколько</span>'
                '<input type="file" accept="image/*" multiple hidden data-photo-upload="gallery">'
                '</label>'
            )
        )

        return format_html(
            '<div class="product-row-detail__gallery" data-photo-gallery data-photos-url="{}">{}</div>',
            photos_url,
            mark_safe(''.join(str(card) for card in cards)),
        )

    def row_details_html(self, obj):
        quick_update_url = reverse('admin:catalog_product_quick_update', args=[obj.pk])
        change_url = reverse('admin:catalog_product_change', args=[obj.pk])
        site_url = obj.get_absolute_url()

        status_options = format_html_join(
            '',
            '<option value="{}"{}>{}</option>',
            (
                (value, ' selected' if value == obj.status else '', label)
                for value, label in Product.Status.choices
            ),
        )

        return format_html(
            '<div class="product-row-detail__inner">'
            '<div class="product-row-detail__notice" data-quick-message></div>'
            '<div class="product-row-detail__block">'
            '<span class="product-row-detail__label">Название</span>'
            '<input class="vTextField product-row-detail__input product-row-detail__name" '
            'data-quick-field="name" value="{}">'
            '</div>'
            '<div class="product-row-detail__grid">'
            '<div><span class="product-row-detail__label">Код</span><input class="vTextField product-row-detail__input" data-quick-field="sku" value="{}"></div>'
            '<div><span class="product-row-detail__label">Штрихкод</span><input class="vTextField product-row-detail__input" data-quick-field="barcode" value="{}"></div>'
            '<div><span class="product-row-detail__label">Слаг</span><input class="vTextField product-row-detail__input" data-quick-field="slug" value="{}"></div>'
            '<div><span class="product-row-detail__label">Ед. изм.</span><input class="vTextField product-row-detail__input" data-quick-field="unit" value="{}"></div>'
            '<div><span class="product-row-detail__label">Страна</span><input class="vTextField product-row-detail__input" data-quick-field="country" value="{}"></div>'
            '<div><span class="product-row-detail__label">Статус</span><select class="product-row-detail__input" data-quick-field="status">{}</select></div>'
            '<div><span class="product-row-detail__label">Рекомендация</span>'
            '<input class="vTextField product-row-detail__input" data-quick-field="recommendation_codes" '
            'value="{}" placeholder="1 или 1,3,5" title="Номера групп через запятую"></div>'
            '<div><span class="product-row-detail__label">Рейтинг</span> {} ({} оценок)</div>'
            '<div><span class="product-row-detail__label">Создан</span> {}</div>'
            '</div>'
            '<div class="product-row-detail__block">'
            '<span class="product-row-detail__label">Описание</span>'
            '<textarea class="vLargeTextField product-row-detail__input product-row-detail__textarea" '
            'data-quick-field="description">{}</textarea>'
            '</div>'
            '<div class="product-row-detail__block">'
            '<span class="product-row-detail__label">Фото</span>'
            '{}'
            '</div>'
            '<div class="product-row-detail__actions">'
            '<button type="button" class="button default product-row-save product-row-detail__save" '
            'data-product-save="{}" data-quick-url="{}" data-quick-save disabled>'
            'Сохранить в строке</button>'
            '<a class="button" href="{}" target="_blank" rel="noopener noreferrer">На сайте</a>'
            '<a class="button" href="{}">Полное редактирование</a>'
            '</div>'
            '</div>',
            obj.name or '',
            obj.sku or '',
            obj.barcode or '',
            obj.slug or '',
            obj.unit or '',
            obj.country or '',
            status_options,
            obj.recommendation_codes or '',
            obj.rating,
            obj.reviews_count,
            obj.created_at.strftime('%d.%m.%Y %H:%M') if obj.created_at else '—',
            obj.description or '',
            self.photos_html(obj),
            obj.pk,
            quick_update_url,
            site_url,
            change_url,
        )

    @admin.display(description='Главное фото', boolean=True)
    def has_main_photo(self, obj):
        return bool(obj.image)

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                '<int:product_id>/quick-update/',
                self.admin_site.admin_view(self.quick_update_view),
                name='catalog_product_quick_update',
            ),
            path(
                '<int:product_id>/quick-photos/',
                self.admin_site.admin_view(self.quick_photos_view),
                name='catalog_product_quick_photos',
            ),
            path(
                '<int:product_id>/quick-details/',
                self.admin_site.admin_view(self.quick_details_view),
                name='catalog_product_quick_details',
            ),
            path(
                'search-suggest/',
                self.admin_site.admin_view(self.search_suggest_view),
                name='catalog_product_search_suggest',
            ),
            path(
                'import/',
                self.admin_site.admin_view(self.import_view),
                name='catalog_product_import',
            ),
        ]
        return custom + urls

    def search_suggest_view(self, request):
        """Подсказки поиска в админке — как на сайте."""
        if not (self.has_view_permission(request) or self.has_change_permission(request)):
            return JsonResponse({'results': []}, status=403)

        q = (request.GET.get('q') or '').strip()
        if len(q) < 1:
            return JsonResponse({'results': []})

        qs = self.get_queryset(request)
        qs = filter_products_by_query(qs, q, prefix_only=True)
        ranked = rank_prefix_first(list(qs[:200]), q)[:12]
        results = [
            {
                'id': p.id,
                'name': p.name,
                'sku': p.sku or '',
                'category': p.category.name if p.category_id else '',
                'url': reverse('admin:catalog_product_change', args=[p.pk]),
            }
            for p in ranked
        ]
        return JsonResponse({'results': results})

    def _ensure_product_card_image(self, product):
        from .image_utils import ensure_card_image, persist_card_image

        product.refresh_from_db()
        if ensure_card_image(product, force=True):
            persist_card_image(product)

    def _product_photos_response(self, product, message='Готово'):
        product.refresh_from_db()
        icon = (
            '<img src="/static/admin/img/icon-yes.svg" alt="True">'
            if product.image
            else '<img src="/static/admin/img/icon-no.svg" alt="False">'
        )
        return JsonResponse({
            'ok': True,
            'message': message,
            'photos_html': str(self.photos_html(product)),
            'has_main_photo_html': icon,
        })

    def quick_details_view(self, request, product_id):
        """HTML раскрытой строки — только по запросу (не на всю страницу списка)."""
        if request.method != 'GET':
            return JsonResponse({'ok': False, 'error': 'GET only'}, status=405)
        if not (self.has_view_permission(request) or self.has_change_permission(request)):
            return JsonResponse({'ok': False, 'error': 'Forbidden'}, status=403)

        product = (
            Product.objects.filter(pk=product_id)
            .select_related('category')
            .prefetch_related('images')
            .first()
        )
        if not product:
            return JsonResponse({'ok': False, 'error': 'Not found'}, status=404)

        return JsonResponse({
            'ok': True,
            'html': str(self.row_details_html(product)),
        })

    def quick_update_view(self, request, product_id):
        if request.method != 'POST':
            return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)
        if not self.has_change_permission(request):
            return JsonResponse({'ok': False, 'error': 'Forbidden'}, status=403)

        try:
            payload = json.loads(request.body.decode('utf-8') or '{}')
        except (TypeError, ValueError, json.JSONDecodeError):
            return JsonResponse({'ok': False, 'error': 'Bad JSON'}, status=400)

        product = Product.objects.filter(pk=product_id).first()
        if not product:
            return JsonResponse({'ok': False, 'error': 'Not found'}, status=404)

        changed_fields = []
        text_fields = (
            'name', 'sku', 'barcode', 'slug', 'unit', 'country',
            'description', 'recommendation_codes',
        )
        for field in text_fields:
            if field in payload:
                value = str(payload.get(field) or '').strip()
                if field == 'name' and not value:
                    return JsonResponse({'ok': False, 'error': 'Название не может быть пустым'}, status=400)
                if field == 'recommendation_codes':
                    from .models import normalize_recommendation_codes

                    value = normalize_recommendation_codes(value)
                setattr(product, field, value)
                changed_fields.append(field)

        if 'status' in payload:
            status = str(payload.get('status') or '').strip()
            valid_statuses = {choice[0] for choice in Product.Status.choices}
            if status in valid_statuses:
                product.status = status
                changed_fields.append('status')

        if 'category' in payload:
            from catalog.models import Category

            raw_cat = payload.get('category')
            try:
                cat_id = int(raw_cat)
            except (TypeError, ValueError):
                return JsonResponse({'ok': False, 'error': 'Некорректная категория'}, status=400)
            if not Category.objects.filter(pk=cat_id).exists():
                return JsonResponse({'ok': False, 'error': 'Категория не найдена'}, status=400)
            product.category_id = cat_id
            changed_fields.append('category')

        if 'price' in payload:
            from decimal import Decimal, InvalidOperation

            raw = str(payload.get('price') or '').replace('\u00a0', '').replace(' ', '').replace(',', '.')
            try:
                product.price = Decimal(raw)
            except (InvalidOperation, TypeError, ValueError):
                return JsonResponse({'ok': False, 'error': 'Некорректная цена'}, status=400)
            changed_fields.append('price')

        if 'old_price' in payload:
            from decimal import Decimal, InvalidOperation

            raw = str(payload.get('old_price') or '').replace('\u00a0', '').replace(' ', '').replace(',', '.')
            if not raw:
                product.old_price = None
            else:
                try:
                    product.old_price = Decimal(raw)
                except (InvalidOperation, TypeError, ValueError):
                    return JsonResponse({'ok': False, 'error': 'Некорректная старая цена'}, status=400)
            changed_fields.append('old_price')

        if 'is_visible' in payload:
            raw = payload.get('is_visible')
            product.is_visible = str(raw).lower() in {'1', 'true', 'yes', 'on'}
            changed_fields.append('is_visible')

        if 'is_featured' in payload:
            raw = payload.get('is_featured')
            product.is_featured = str(raw).lower() in {'1', 'true', 'yes', 'on'}
            changed_fields.append('is_featured')

        if not changed_fields:
            return JsonResponse({'ok': True, 'message': 'Нет изменений', 'slug': product.slug})

        product.save()
        return JsonResponse({
            'ok': True,
            'message': 'Сохранено — изменения уже на сайте',
            'slug': product.slug,
        })

    def quick_photos_view(self, request, product_id):
        try:
            return self._quick_photos_view(request, product_id)
        except Exception:
            logger.exception('Ошибка загрузки фото товара id=%s', product_id)
            return JsonResponse(
                {'ok': False, 'error': 'Не удалось сохранить фото. Попробуйте ещё раз.'},
                status=500,
            )

    def _quick_photos_view(self, request, product_id):
        if request.method != 'POST':
            return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)
        if not self.has_change_permission(request):
            return JsonResponse({'ok': False, 'error': 'Forbidden'}, status=403)

        product = (
            Product.objects.filter(pk=product_id)
            .prefetch_related('images')
            .first()
        )
        if not product:
            return JsonResponse({'ok': False, 'error': 'Not found'}, status=404)

        action = (request.POST.get('action') or '').strip()

        if action == 'upload_main':
            uploaded = request.FILES.get('image')
            if not uploaded:
                return JsonResponse({'ok': False, 'error': 'Файл не выбран'}, status=400)
            product.image = uploaded
            product.save()
            self._ensure_product_card_image(product)
            return self._product_photos_response(product, 'Главное фото обновлено')

        if action == 'upload_gallery':
            files = request.FILES.getlist('images') or (
                [request.FILES['image']] if request.FILES.get('image') else []
            )
            if not files:
                return JsonResponse({'ok': False, 'error': 'Файлы не выбраны'}, status=400)
            next_order = (
                ProductImage.objects.filter(product=product)
                .order_by('-sort_order')
                .values_list('sort_order', flat=True)
                .first()
            )
            order = 0 if next_order is None else next_order + 1
            for uploaded in files:
                ProductImage.objects.create(product=product, image=uploaded, sort_order=order)
                order += 1
            return self._product_photos_response(product, f'Добавлено фото: {len(files)}')

        if action == 'delete_main':
            if product.image:
                product.image.delete(save=False)
                product.image = None
                product.save()
            return self._product_photos_response(product, 'Главное фото удалено')

        if action == 'delete_gallery':
            try:
                image_id = int(request.POST.get('image_id') or 0)
            except (TypeError, ValueError):
                image_id = 0
            image = ProductImage.objects.filter(product=product, pk=image_id).first()
            if not image:
                return JsonResponse({'ok': False, 'error': 'Фото не найдено'}, status=404)
            image.image.delete(save=False)
            image.delete()
            return self._product_photos_response(product, 'Фото удалено')

        if action == 'set_main':
            try:
                image_id = int(request.POST.get('image_id') or 0)
            except (TypeError, ValueError):
                image_id = 0
            image = ProductImage.objects.filter(product=product, pk=image_id).first()
            if not image or not image.image:
                return JsonResponse({'ok': False, 'error': 'Фото не найдено'}, status=404)

            old_main_name = None
            old_main_content = None
            if product.image:
                product.image.open('rb')
                try:
                    old_main_content = product.image.read()
                    old_main_name = Path(product.image.name).name
                finally:
                    product.image.close()

            image.image.open('rb')
            try:
                new_content = image.image.read()
                new_name = Path(image.image.name).name
            finally:
                image.image.close()

            product.image.save(new_name, ContentFile(new_content), save=True)
            self._ensure_product_card_image(product)
            image.image.delete(save=False)
            image.delete()

            if old_main_content and old_main_name:
                next_order = (
                    ProductImage.objects.filter(product=product)
                    .order_by('-sort_order')
                    .values_list('sort_order', flat=True)
                    .first()
                )
                order = 0 if next_order is None else next_order + 1
                gallery = ProductImage(product=product, sort_order=order)
                gallery.image.save(old_main_name, ContentFile(old_main_content), save=True)

            return self._product_photos_response(product, 'Главное фото заменено')

        return JsonResponse({'ok': False, 'error': 'Unknown action'}, status=400)

    def import_view(self, request):
        if request.method == 'POST' and request.FILES.get('file'):
            uploaded = request.FILES['file']
            filename = uploaded.name.lower()
            try:
                if filename.endswith('.csv'):
                    created, updated, errors = self._import_csv(uploaded)
                elif filename.endswith('.xlsx'):
                    created, updated, errors = self._import_xlsx(uploaded)
                elif filename.endswith('.xls'):
                    created, updated, errors = self._import_xls(uploaded)
                else:
                    messages.error(request, 'Поддерживаются файлы CSV, XLS и XLSX.')
                    return redirect('admin:catalog_product_import')
            except Exception as exc:  # noqa: BLE001
                messages.error(request, f'Ошибка импорта: {exc}')
                return redirect('admin:catalog_product_import')

            messages.success(
                request,
                f'Импорт завершён: создано {created}, обновлено {updated}. '
                'Товары автоматически распределены по категориям.',
            )
            for err in errors[:30]:
                messages.warning(request, err)
            return redirect('admin:catalog_product_changelist')

        context = {
            **self.admin_site.each_context(request),
            'title': 'Массовая загрузка товаров из Excel',
            'opts': self.model._meta,
        }
        return render(request, 'admin/catalog/product/import.html', context)

    def _parse_row(self, row, default_category=None):
        name = (row.get('name') or '').strip()
        if not name:
            raise ValueError('Пустое наименование товара')

        description = (row.get('description') or '').strip()
        # Старые файлы могли иметь «краткое описание» — дописываем в полное, если описания нет
        legacy_short = (row.get('short_description') or '').strip()
        if legacy_short and not description:
            description = legacy_short
        elif legacy_short and legacy_short not in description:
            description = f'{legacy_short}\n{description}'.strip()

        explicit_category = (row.get('category') or '').strip()
        category = resolve_category(name, description, explicit_category)

        try:
            price = parse_price(row.get('price'))
        except InvalidOperation as exc:
            raise ValueError(f'Некорректная цена для «{name}»') from exc
        if price is None:
            raise ValueError(f'Нет цены для «{name}»')

        old_price = None
        try:
            old_price = parse_price(row.get('old_price'))
        except InvalidOperation:
            old_price = None

        status_raw = (row.get('status') or 'in_stock').strip().lower()
        status_map = {
            'in_stock': Product.Status.IN_STOCK,
            'в наличии': Product.Status.IN_STOCK,
            'in_transit': Product.Status.IN_TRANSIT,
            'в пути': Product.Status.IN_TRANSIT,
            'out_of_stock': Product.Status.OUT_OF_STOCK,
            'нет в наличии': Product.Status.OUT_OF_STOCK,
            'on_order': Product.Status.ON_ORDER,
            'под заказ': Product.Status.ON_ORDER,
        }
        status = status_map.get(status_raw, Product.Status.IN_STOCK)

        promo_raw = str(row.get('is_promo', '')).strip().lower()
        is_promo = promo_raw in {'1', 'true', 'yes', 'да', 'y'}
        if not is_promo and old_price and old_price > price:
            is_promo = True

        return {
            'name': name,
            'category': category,
            'price': price,
            'old_price': old_price,
            'description': description,
            'unit': '',
            'country': (row.get('country') or '').strip(),
            'sku': '',
            'barcode': '',
            'status': status,
            'is_promo': is_promo,
            'is_visible': True,
        }

    def _upsert_product(self, data):
        product = None
        if data.get('sku'):
            product = Product.objects.filter(sku=data['sku']).first()
        if not product:
            product = Product.objects.filter(name=data['name']).first()
        if product:
            for key, value in data.items():
                setattr(product, key, value)
            product.save()
            return False
        Product.objects.create(**data)
        return True

    def _import_mapped_rows(self, headers, data_rows, default_category=None, start_row=2):
        mapping = map_headers(headers)
        if 'name' not in mapping or 'price' not in mapping:
            raise ValueError(
                'В файле должны быть столбцы «Наименование товара» и «Цена». '
                f'Найдены: {headers}'
            )
        created = updated = 0
        errors = []
        for offset, values in enumerate(data_rows):
            row_number = start_row + offset
            if not values or all(v is None or str(v).strip() == '' for v in values):
                continue
            row = {}
            for field, idx in mapping.items():
                if idx < len(values):
                    value = values[idx]
                    row[field] = '' if value is None else value
            try:
                data = self._parse_row(row)
                if self._upsert_product(data):
                    created += 1
                else:
                    updated += 1
            except Exception as exc:  # noqa: BLE001
                errors.append(f'Строка {row_number}: {exc}')
        return created, updated, errors

    def _import_csv(self, uploaded, default_category=None):
        content = uploaded.read().decode('utf-8-sig')
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        if not rows:
            return 0, 0, ['Пустой файл']
        return self._import_mapped_rows(rows[0], rows[1:])

    def _import_xlsx(self, uploaded, default_category=None):
        wb = load_workbook(uploaded, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return 0, 0, ['Пустой файл']
        return self._import_mapped_rows(rows[0], rows[1:])

    def _import_xls(self, uploaded, default_category=None):
        import xlrd

        book = xlrd.open_workbook(file_contents=uploaded.read())
        sheet = book.sheet_by_index(0)
        if sheet.nrows < 2:
            return 0, 0, ['Пустой файл']
        headers = [sheet.cell_value(0, c) for c in range(sheet.ncols)]
        data_rows = [
            [sheet.cell_value(r, c) for c in range(sheet.ncols)]
            for r in range(1, sheet.nrows)
        ]
        return self._import_mapped_rows(headers, data_rows)


@admin.register(ProductImage)
class ProductImageAdmin(admin.ModelAdmin):
    list_display = ('thumb', 'product', 'sort_order', 'image')
    list_editable = ('sort_order',)
    list_filter = ('product__category',)
    search_fields = ('product__name', 'product__sku', 'product__barcode', 'image')
    autocomplete_fields = ['product']
    list_select_related = ('product',)
    list_per_page = 50
    change_list_template = 'admin/catalog/productimage/change_list.html'

    @admin.display(description='Превью')
    def thumb(self, obj):
        if obj.image:
            return format_html(
                '<img src="{}" style="height:48px;width:48px;object-fit:cover;border-radius:6px;" />',
                obj.image.url,
            )
        return '—'

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                'bulk-upload/',
                self.admin_site.admin_view(self.bulk_upload_view),
                name='catalog_productimage_bulk_upload',
            ),
        ]
        return custom + urls

    def bulk_upload_view(self, request):
        from .photo_import import assign_photos, iter_upload_files

        if request.method == 'POST':
            zip_file = request.FILES.get('zip_file')
            files = request.FILES.getlist('images') + request.FILES.getlist('folder_images')
            replace_main = request.POST.get('replace_main') == '1'

            if not zip_file and not files:
                messages.error(request, 'Выберите ZIP-архив или файлы/папку с фото.')
                return redirect('admin:catalog_productimage_bulk_upload')

            try:
                items = list(iter_upload_files(files, zip_file=zip_file))
            except Exception as exc:  # noqa: BLE001
                messages.error(request, f'Не удалось прочитать файлы: {exc}')
                return redirect('admin:catalog_productimage_bulk_upload')

            if not items:
                messages.error(
                    request,
                    'Не найдено изображений. Поддерживаются JPG, PNG, WEBP, GIF, BMP, TIFF.',
                )
                return redirect('admin:catalog_productimage_bulk_upload')

            result = assign_photos(items, replace_main=replace_main)
            messages.success(
                request,
                f'Готово: главных фото — {result.assigned_main}, '
                f'в галерею — {result.assigned_gallery}.',
            )
            for name in result.unmatched[:40]:
                messages.warning(request, f'Не найден товар для файла: {name}')
            if len(result.unmatched) > 40:
                messages.warning(
                    request,
                    f'…и ещё {len(result.unmatched) - 40} файлов без совпадения.',
                )
            for err in result.errors[:20]:
                messages.error(request, err)
            for skip in result.skipped[:10]:
                messages.info(request, skip)
            return redirect('admin:catalog_productimage_changelist')

        context = {
            **self.admin_site.each_context(request),
            'title': 'Массовая загрузка фото',
            'opts': self.model._meta,
        }
        return render(request, 'admin/catalog/productimage/bulk_upload.html', context)


@admin.register(ProductRecommendation)
class ProductRecommendationAdmin(admin.ModelAdmin):
    list_display = ('product', 'recommended_product', 'sort_order')
    autocomplete_fields = ['product', 'recommended_product']


@admin.register(ProductReview)
class ProductReviewAdmin(admin.ModelAdmin):
    list_display = ('product', 'user', 'rating', 'created_at')
    list_filter = ('rating', 'created_at')
    search_fields = ('product__name', 'user__username', 'comment')
    autocomplete_fields = ['product', 'user']
    readonly_fields = ('created_at', 'updated_at')
