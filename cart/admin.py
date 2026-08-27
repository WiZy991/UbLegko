import io
import json
from urllib.parse import quote

from django.contrib import admin, messages
from django.contrib.admin.templatetags.admin_list import (
    ResultList,
    items_for_result,
    result_headers,
    result_hidden_fields,
)
from django.http import HttpResponse, JsonResponse
from django.urls import path, reverse
from django.utils import timezone
from django.utils.html import escape, format_html, mark_safe
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from core.models import SiteSettings

from .models import Favorite, Order, OrderItem, StainHelpRequest


def format_local_datetime(value):
    if not value:
        return '—'
    return timezone.localtime(value).strftime('%d.%m.%Y %H:%M')


def inbox_status_buttons(url, current, kind, done_label):
    """Две кнопки статуса с мгновенным сохранением."""
    new_active = ' is-active' if current == 'new' else ''
    done_active = ' is-active' if current == 'processed' else ''
    return format_html(
        '<div class="inbox-status-btns" data-inbox-quick-url="{}" '
        'data-inbox-kind="{}" data-inbox-status="{}" role="group" aria-label="Статус">'
        '<button type="button" class="inbox-status-btn inbox-status-btn--new{}" '
        'data-status="new">Новая</button>'
        '<button type="button" class="inbox-status-btn inbox-status-btn--done{}" '
        'data-status="processed">{}</button>'
        '</div>',
        url,
        kind,
        current,
        new_active,
        done_active,
        done_label,
    )


class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 0
    readonly_fields = ('product', 'product_name', 'price', 'quantity')
    can_delete = False


class InboxExpandableAdminMixin:
    """Раскрывающиеся строки + счётчик новых для changelist."""

    change_list_template = None  # задаётся в наследниках

    def get_inbox_new_count(self):
        raise NotImplementedError

    def get_row_status(self, obj):
        """Значение data-inbox-status: new | processed."""
        raise NotImplementedError

    def row_details_html(self, obj):
        raise NotImplementedError

    @admin.display(description='')
    def expand_toggle(self, obj):
        return format_html(
            '<button type="button" class="inbox-row-expand" data-inbox-id="{}" '
            'aria-expanded="false" aria-label="Подробнее" title="Подробнее">'
            '<i class="fas fa-chevron-down" aria-hidden="true"></i>'
            '</button>',
            obj.pk,
        )

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        new_count = self.get_inbox_new_count()
        extra_context['new_count'] = new_count
        plural = self.model._meta.verbose_name_plural
        extra_context['title'] = f'{plural} ({new_count})'
        response = super().changelist_view(request, extra_context=extra_context)
        if not hasattr(response, 'context_data'):
            return response

        cl = response.context_data.get('cl')
        if not cl:
            return response

        rows = []
        if cl.formset:
            for res, form in zip(cl.result_list, cl.formset.forms):
                rows.append({
                    'result': ResultList(form, items_for_result(cl, res, form)),
                    'obj': res,
                    'row_status': self.get_row_status(res),
                    'details_html': self.row_details_html(res),
                })
        else:
            for res in cl.result_list:
                rows.append({
                    'result': ResultList(None, items_for_result(cl, res, None)),
                    'obj': res,
                    'row_status': self.get_row_status(res),
                    'details_html': self.row_details_html(res),
                })

        headers = list(result_headers(cl))
        num_sorted_fields = sum(
            1 for header in headers if header.get('sortable') and header.get('sorted')
        )
        response.context_data.update({
            'expandable_rows': rows,
            'result_headers': headers,
            'num_sorted_fields': num_sorted_fields,
            'result_hidden_fields': list(result_hidden_fields(cl)),
            'new_count': new_count,
        })
        return response


@admin.register(Order)
class OrderAdmin(InboxExpandableAdminMixin, admin.ModelAdmin):
    change_list_template = 'admin/cart/order/change_list.html'
    list_display = (
        'expand_toggle',
        'id',
        'full_name',
        'phone',
        'city',
        'delivery_col',
        'email',
        'email_sent_col',
        'address_col',
        'address_name_col',
        'status_control',
        'created_col',
        'total_display',
    )
    list_filter = ('status', 'delivery_method', 'email_sent', 'created_at')
    search_fields = ('full_name', 'phone', 'email', 'address', 'address_name', 'city')
    readonly_fields = (
        'user',
        'full_name',
        'phone',
        'email',
        'email_sent',
        'delivery_method',
        'address',
        'address_name',
        'city',
        'comment',
        'site_feedback',
        'created_at',
    )
    inlines = [OrderItemInline]
    list_per_page = 50
    actions = (
        'export_delivery_sheet',
        'export_orders_sheet',
        'delete_selected',
    )

    @admin.action(description='Сформировать лист доставки')
    def export_delivery_sheet(self, request, queryset):
        orders = list(queryset.order_by('created_at', 'id'))
        if not orders:
            self.message_user(request, 'Не выбрано ни одной заявки', level=messages.WARNING)
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = 'Лист доставки'
        headers = ('Имя', 'Телефон', 'Адрес доставки', 'Комментарий')
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical='center', wrap_text=True)

        for order in orders:
            address = self._delivery_address_for_sheet(order)
            comment = (order.comment or '').strip()
            row = (
                (order.full_name or '').strip(),
                (order.phone or '').strip(),
                address,
                comment,
            )
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 42
        ws.column_dimensions['D'].width = 40
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:D{ws.max_row}'

        return self._xlsx_response(
            wb,
            download_name=f'Лист доставки на {timezone.localtime().strftime("%d.%m.%y")}.xlsx',
            ascii_fallback=f'delivery_list_{timezone.localtime().strftime("%d.%m.%y")}.xlsx',
        )

    @admin.action(description='Сформировать лист заявок')
    def export_orders_sheet(self, request, queryset):
        orders = list(
            queryset.order_by('created_at', 'id').prefetch_related('items')
        )
        if not orders:
            self.message_user(request, 'Не выбрано ни одной заявки', level=messages.WARNING)
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = 'Лист заявок'
        headers = (
            'Имя',
            'Телефон',
            'Адрес',
            'Товары',
            'Количество',
            'Общая сумма',
        )
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical='center', wrap_text=True)

        for order in orders:
            items = list(order.items.all())
            products = '\n'.join(
                (item.product_name or '').strip() or '—' for item in items
            ) or '—'
            quantities = '\n'.join(str(item.quantity) for item in items) or '0'
            row = (
                (order.full_name or '').strip(),
                (order.phone or '').strip(),
                self._delivery_address_for_sheet(order),
                products,
                quantities,
                f'{order.total:.0f} руб',
            )
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 42
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:F{ws.max_row}'

        return self._xlsx_response(
            wb,
            download_name=f'Лист заявок на {timezone.localtime().strftime("%d.%m.%y")}.xlsx',
            ascii_fallback=f'orders_list_{timezone.localtime().strftime("%d.%m.%y")}.xlsx',
        )

    @staticmethod
    def _xlsx_response(wb, *, download_name, ascii_fallback):
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type=(
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            ),
        )
        response['Content-Disposition'] = (
            f'attachment; filename="{ascii_fallback}"; '
            f"filename*=UTF-8''{quote(download_name)}"
        )
        return response

    @staticmethod
    def _delivery_address_for_sheet(order):
        if order.delivery_method == 'pickup' and not (order.address or '').strip():
            return 'Самовывоз'
        parts = []
        if (order.city or '').strip():
            parts.append(order.city.strip())
        if (order.address or '').strip():
            parts.append(order.address.strip())
        address = ', '.join(parts)
        name = (order.address_name or '').strip()
        if name and address:
            return f'{name}: {address}'
        return name or address or '—'

    def get_queryset(self, request):
        return super().get_queryset(request).prefetch_related('items')

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                'inbox-counts/',
                self.admin_site.admin_view(self.inbox_counts_view),
                name='cart_inbox_counts',
            ),
            path(
                '<path:object_id>/quick-status/',
                self.admin_site.admin_view(self.quick_status_view),
                name='cart_order_quick_status',
            ),
        ]
        return custom + urls

    def inbox_counts_view(self, request):
        return JsonResponse({
            'orders_new': Order.objects.filter(status=Order.Status.NEW).count(),
            'requests_new': StainHelpRequest.objects.filter(is_processed=False).count(),
        })

    def quick_status_view(self, request, object_id):
        if request.method != 'POST':
            return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)
        if not self.has_change_permission(request):
            return JsonResponse({'ok': False, 'error': 'Forbidden'}, status=403)

        try:
            payload = json.loads(request.body.decode('utf-8') or '{}')
        except (TypeError, ValueError, json.JSONDecodeError):
            return JsonResponse({'ok': False, 'error': 'Bad JSON'}, status=400)

        obj = Order.objects.filter(pk=object_id).first()
        if not obj:
            return JsonResponse({'ok': False, 'error': 'Not found'}, status=404)

        status = str(payload.get('status') or '').strip()
        valid = {choice[0] for choice in Order.Status.choices}
        if status not in valid:
            return JsonResponse({'ok': False, 'error': 'Некорректный статус'}, status=400)

        if obj.status != status:
            obj.status = status
            obj.save(update_fields=['status'])

        return JsonResponse({
            'ok': True,
            'status': obj.status,
            'orders_new': Order.objects.filter(status=Order.Status.NEW).count(),
            'requests_new': StainHelpRequest.objects.filter(is_processed=False).count(),
        })

    def get_inbox_new_count(self):
        return Order.objects.filter(status=Order.Status.NEW).count()

    def get_row_status(self, obj):
        return obj.status if obj.status in ('new', 'processed') else 'new'

    @admin.display(description='Получение', ordering='delivery_method')
    def delivery_col(self, obj):
        return obj.get_delivery_method_display()

    @admin.display(description='Письмо', boolean=True, ordering='email_sent')
    def email_sent_col(self, obj):
        return obj.email_sent

    @admin.display(description='Адрес', ordering='address')
    def address_col(self, obj):
        return obj.address

    @admin.display(description='Название', ordering='address_name')
    def address_name_col(self, obj):
        return obj.address_name

    @admin.display(description='Статус', ordering='status')
    def status_control(self, obj):
        return inbox_status_buttons(
            reverse('admin:cart_order_quick_status', args=[obj.pk]),
            obj.status if obj.status in ('new', 'processed') else 'new',
            'order',
            'Готово',
        )

    @admin.display(description='Создана', ordering='created_at')
    def created_col(self, obj):
        return format_local_datetime(obj.created_at)

    @admin.display(description='Сумма')
    def total_display(self, obj):
        return f'{obj.total:.0f} руб'

    def row_details_html(self, obj):
        items = list(obj.items.all())
        if items:
            items_html = '<ul class="inbox-row-detail__items">' + ''.join(
                f'<li>{escape(item.product_name)} × {item.quantity} — '
                f'{item.line_total:.0f} руб</li>'
                for item in items
            ) + (
                f'<li class="inbox-row-detail__total"><strong>Итого: '
                f'{obj.total:.0f} руб</strong></li>'
            ) + '</ul>'
        else:
            items_html = '<span class="inbox-row-detail__value">Нет позиций</span>'

        comment = (obj.comment or '').strip() or '—'
        feedback = (obj.site_feedback or '').strip() or '—'
        address = (obj.address or '').strip() or '—'
        address_name = (obj.address_name or '').strip() or '—'
        created = format_local_datetime(obj.created_at)
        email = (obj.email or '').strip() or '—'
        delivery = obj.get_delivery_method_display() if obj.delivery_method else '—'
        email_sent = 'Да' if obj.email_sent else 'Нет'

        return mark_safe(
            '<div class="inbox-row-detail__inner">'
            '<div class="inbox-row-detail__grid">'
            f'<div class="inbox-row-detail__block">'
            f'<span class="inbox-row-detail__label">Email</span>'
            f'<div class="inbox-row-detail__value">{escape(email)}</div></div>'
            f'<div class="inbox-row-detail__block">'
            f'<span class="inbox-row-detail__label">Когда создана</span>'
            f'<div class="inbox-row-detail__value">{escape(created)}</div></div>'
            f'<div class="inbox-row-detail__block">'
            f'<span class="inbox-row-detail__label">Доставка</span>'
            f'<div class="inbox-row-detail__value">{escape(str(delivery))}</div></div>'
            f'<div class="inbox-row-detail__block">'
            f'<span class="inbox-row-detail__label">Письмо отправлено</span>'
            f'<div class="inbox-row-detail__value">{escape(email_sent)}</div></div>'
            f'<div class="inbox-row-detail__block">'
            f'<span class="inbox-row-detail__label">Название адреса</span>'
            f'<div class="inbox-row-detail__value">{escape(address_name)}</div></div>'
            f'<div class="inbox-row-detail__block inbox-row-detail__block--wide">'
            f'<span class="inbox-row-detail__label">Адрес</span>'
            f'<div class="inbox-row-detail__value">{escape(address)}</div></div>'
            f'<div class="inbox-row-detail__block inbox-row-detail__block--wide">'
            f'<span class="inbox-row-detail__label">Комментарий</span>'
            f'<div class="inbox-row-detail__value">{escape(comment)}</div></div>'
            f'<div class="inbox-row-detail__block inbox-row-detail__block--wide">'
            f'<span class="inbox-row-detail__label">Замечания по сайту</span>'
            f'<div class="inbox-row-detail__value">{escape(feedback)}</div></div>'
            f'<div class="inbox-row-detail__block inbox-row-detail__block--wide">'
            f'<span class="inbox-row-detail__label">Состав заказа</span>'
            f'{items_html}</div>'
            '</div></div>'
        )


@admin.register(Favorite)
class FavoriteAdmin(admin.ModelAdmin):
    list_display = ('user', 'product', 'created_at')
    search_fields = ('user__username', 'product__name')


@admin.register(StainHelpRequest)
class StainHelpRequestAdmin(InboxExpandableAdminMixin, admin.ModelAdmin):
    change_list_template = 'admin/cart/stainhelprequest/change_list.html'
    list_display = (
        'expand_toggle',
        'request_number',
        'created_col',
        'full_name',
        'phone',
        'contact_col',
        'short_problem',
        'status_control',
    )
    list_filter = ('is_processed', 'email_sent', 'created_at')
    search_fields = ('full_name', 'phone', 'contact_method', 'problem')
    readonly_fields = (
        'full_name',
        'phone',
        'contact_method',
        'problem',
        'user',
        'email_sent',
        'created_at',
    )
    fields = (
        'created_at',
        'full_name',
        'phone',
        'contact_method',
        'problem',
        'user',
        'email_sent',
        'is_processed',
    )
    ordering = ('-created_at',)
    date_hierarchy = 'created_at'
    list_per_page = 50
    actions = ('export_requests_sheet', 'resend_email', 'delete_selected')

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                'auto-modal/',
                self.admin_site.admin_view(self.auto_modal_view),
                name='cart_stainhelprequest_auto_modal',
            ),
            path(
                '<path:object_id>/quick-status/',
                self.admin_site.admin_view(self.quick_status_view),
                name='cart_stainhelprequest_quick_status',
            ),
        ]
        return custom + urls

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['stain_help_auto_modal'] = bool(
            SiteSettings.load().stain_help_auto_modal
        )
        extra_context['stain_help_auto_modal_url'] = reverse(
            'admin:cart_stainhelprequest_auto_modal'
        )
        extra_context['can_toggle_stain_help_auto_modal'] = self.has_change_permission(
            request
        )
        return super().changelist_view(request, extra_context)

    def auto_modal_view(self, request):
        if request.method != 'POST':
            return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)
        if not self.has_change_permission(request):
            return JsonResponse({'ok': False, 'error': 'Forbidden'}, status=403)

        try:
            payload = json.loads(request.body.decode('utf-8') or '{}')
        except (TypeError, ValueError, json.JSONDecodeError):
            return JsonResponse({'ok': False, 'error': 'Bad JSON'}, status=400)

        raw = payload.get('enabled')
        if isinstance(raw, str):
            enabled = raw.strip().lower() in {'1', 'true', 'on', 'yes'}
        else:
            enabled = bool(raw)

        site = SiteSettings.load()
        if site.stain_help_auto_modal != enabled:
            site.stain_help_auto_modal = enabled
            site.save(update_fields=['stain_help_auto_modal'])

        return JsonResponse({'ok': True, 'enabled': site.stain_help_auto_modal})

    def quick_status_view(self, request, object_id):
        if request.method != 'POST':
            return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)
        if not self.has_change_permission(request):
            return JsonResponse({'ok': False, 'error': 'Forbidden'}, status=403)

        try:
            payload = json.loads(request.body.decode('utf-8') or '{}')
        except (TypeError, ValueError, json.JSONDecodeError):
            return JsonResponse({'ok': False, 'error': 'Bad JSON'}, status=400)

        obj = StainHelpRequest.objects.filter(pk=object_id).first()
        if not obj:
            return JsonResponse({'ok': False, 'error': 'Not found'}, status=404)

        status = str(payload.get('status') or '').strip()
        if status not in {'new', 'processed'}:
            return JsonResponse({'ok': False, 'error': 'Некорректный статус'}, status=400)

        is_processed = status == 'processed'
        if obj.is_processed != is_processed:
            obj.is_processed = is_processed
            obj.save(update_fields=['is_processed'])

        return JsonResponse({
            'ok': True,
            'status': 'processed' if obj.is_processed else 'new',
            'orders_new': Order.objects.filter(status=Order.Status.NEW).count(),
            'requests_new': StainHelpRequest.objects.filter(is_processed=False).count(),
        })

    def get_inbox_new_count(self):
        return StainHelpRequest.objects.filter(is_processed=False).count()

    def get_row_status(self, obj):
        return 'processed' if obj.is_processed else 'new'

    @admin.display(description='№', ordering='pk')
    def request_number(self, obj):
        return obj.pk

    @admin.display(description='Создано', ordering='created_at')
    def created_col(self, obj):
        return format_local_datetime(obj.created_at)

    @admin.display(description='Связь', ordering='contact_method')
    def contact_col(self, obj):
        return obj.contact_method

    @admin.display(description='Проблема')
    def short_problem(self, obj):
        text = (obj.problem or '').strip().replace('\n', ' ')
        return text if len(text) <= 40 else f'{text[:37]}…'

    @admin.display(description='Статус', ordering='is_processed')
    def status_control(self, obj):
        return inbox_status_buttons(
            reverse('admin:cart_stainhelprequest_quick_status', args=[obj.pk]),
            'processed' if obj.is_processed else 'new',
            'request',
            'Готово',
        )

    def row_details_html(self, obj):
        problem = (obj.problem or '').strip() or '—'
        contact = (obj.contact_method or '').strip() or '—'
        phone = (obj.phone or '').strip() or '—'
        name = (obj.full_name or '').strip() or '—'
        email_sent = 'Да' if obj.email_sent else 'Нет'
        user = str(obj.user) if obj.user_id else '—'
        return mark_safe(
            '<div class="inbox-row-detail__inner">'
            '<div class="inbox-row-detail__grid">'
            f'<div class="inbox-row-detail__block">'
            f'<span class="inbox-row-detail__label">№</span>'
            f'<div class="inbox-row-detail__value">{obj.pk}</div></div>'
            f'<div class="inbox-row-detail__block">'
            f'<span class="inbox-row-detail__label">Имя</span>'
            f'<div class="inbox-row-detail__value">{escape(name)}</div></div>'
            f'<div class="inbox-row-detail__block">'
            f'<span class="inbox-row-detail__label">Телефон</span>'
            f'<div class="inbox-row-detail__value">{escape(phone)}</div></div>'
            f'<div class="inbox-row-detail__block">'
            f'<span class="inbox-row-detail__label">Способ связи</span>'
            f'<div class="inbox-row-detail__value">{escape(contact)}</div></div>'
            f'<div class="inbox-row-detail__block">'
            f'<span class="inbox-row-detail__label">Письмо отправлено</span>'
            f'<div class="inbox-row-detail__value">{escape(email_sent)}</div></div>'
            f'<div class="inbox-row-detail__block">'
            f'<span class="inbox-row-detail__label">Пользователь</span>'
            f'<div class="inbox-row-detail__value">{escape(user)}</div></div>'
            f'<div class="inbox-row-detail__block inbox-row-detail__block--wide">'
            f'<span class="inbox-row-detail__label">Сообщение</span>'
            f'<div class="inbox-row-detail__value">{escape(problem)}</div></div>'
            '</div></div>'
        )

    def has_add_permission(self, request):
        return False

    @admin.action(description='Сформировать лист запросов')
    def export_requests_sheet(self, request, queryset):
        requests_list = list(queryset.order_by('created_at', 'id'))
        if not requests_list:
            self.message_user(request, 'Не выбрано ни одного запроса', level=messages.WARNING)
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = 'Лист запросов'
        headers = ('№', 'Имя', 'Телефон', 'Способ связи', 'Содержание запроса')
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical='center', wrap_text=True)

        for obj in requests_list:
            row = (
                obj.pk,
                (obj.full_name or '').strip(),
                (obj.phone or '').strip(),
                (obj.contact_method or '').strip(),
                (obj.problem or '').strip(),
            )
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 50
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:E{ws.max_row}'

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        date_label = timezone.localtime().strftime('%d.%m.%y')
        download_name = f'Лист запросов на {date_label}.xlsx'
        ascii_fallback = f'requests_list_{timezone.localtime().strftime("%d.%m.%y")}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type=(
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            ),
        )
        response['Content-Disposition'] = (
            f'attachment; filename="{ascii_fallback}"; '
            f"filename*=UTF-8''{quote(download_name)}"
        )
        return response

    @admin.action(description='Отправить письмо на почту магазина ещё раз')
    def resend_email(self, request, queryset):
        from django.template.loader import render_to_string

        from core.mail import send_shop_email

        site = SiteSettings.load()
        ok = 0
        fail = 0
        for obj in queryset:
            subject = (
                f'Запрос №{obj.pk} с сайта {site.brand_name}: '
                f'{obj.full_name}, {obj.phone}'
            )
            body = render_to_string(
                'core/email/stain_help.txt',
                {
                    'site': site,
                    'request_obj': obj,
                    'full_name': obj.full_name,
                    'phone': obj.phone,
                    'contact_method': obj.contact_method,
                    'problem': obj.problem,
                },
            )
            if send_shop_email(subject=subject, body=body, log_label=f'Запрос №{obj.pk}'):
                if not obj.email_sent:
                    obj.email_sent = True
                    obj.save(update_fields=['email_sent'])
                ok += 1
            else:
                fail += 1
        if ok:
            self.message_user(request, f'Отправлено писем: {ok}')
        if fail:
            self.message_user(
                request,
                f'Не удалось отправить: {fail}. Проверьте SMTP в .env и логи.',
                level=30,
            )
